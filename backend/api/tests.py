import re
from datetime import date
from decimal import Decimal
from io import BytesIO
from unittest.mock import patch

from django.contrib.auth.models import User
from django.core import mail
from django.test import override_settings
from django.urls import reverse
from rest_framework import status
from rest_framework.test import APITestCase

from openpyxl import load_workbook

from .angel import AngelAPIError
from .models import (
    AlgoConfiguration,
    EmailOTP,
    Instrument,
    Strategy,
    StrategyActivation,
    Trade,
    UserProfile,
)
from .services.market_data import BaseMarketDataProvider
from .services.strategy_alpha import StrategyAlphaEngine


class StrategyAPITestCase(APITestCase):
    def setUp(self):
        self.list_url = reverse("strategy-list")

    def test_create_strategy(self):
        payload = {
            "name": "Mean Reversion",
            "symbol": "MSFT",
            "timeframe": "1h",
            "status": "draft",
        }
        response = self.client.post(self.list_url, data=payload, format="json")
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(Strategy.objects.count(), 1)
        self.assertEqual(Strategy.objects.first().name, "Mean Reversion")

    def test_list_strategies(self):
        Strategy.objects.create(name="Breakout", symbol="AAPL")
        response = self.client.get(self.list_url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.json()), 1)


class AuthAPITestCase(APITestCase):
    def setUp(self):
        self.request_otp_url = reverse("request-otp")
        self.signup_url = reverse("signup")
        self.login_url = reverse("login")
        self.password_request_url = reverse("password-request-reset")
        self.password_verify_url = reverse("password-verify-otp")
        self.password_reset_url = reverse("password-reset")
        self.admin_users_url = reverse("admin-users")

    def _extract_otp(self, email_body: str) -> str:
        match = re.search(r"(\d{6})", email_body)
        self.assertIsNotNone(match, "OTP not found in email body")
        return match.group(1)

    def test_signup_flow(self):
        email = "new.trader@example.com"
        username = "quantuser"
        otp_response = self.client.post(
            self.request_otp_url,
            {"email": email, "username": username},
            format="json",
        )
        self.assertEqual(otp_response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(mail.outbox), 1)

        otp = self._extract_otp(mail.outbox[0].body)
        payload = {
            "username": username,
            "user_id": "AB1234",
            "api_key": "XYZ123456",
            "mobile": "9876543210",
            "email": email,
            "password": "StrongPass@123",
            "confirm_password": "StrongPass@123",
            "otp": otp,
        }

        signup_response = self.client.post(self.signup_url, payload, format="json")
        self.assertEqual(signup_response.status_code, status.HTTP_201_CREATED)
        self.assertTrue(User.objects.filter(username="quantuser").exists())
        user = User.objects.get(username="quantuser")
        self.assertTrue(UserProfile.objects.filter(user=user).exists())
        instruments = Instrument.objects.filter(user=user).order_by("instrument")
        self.assertEqual(instruments.count(), 3)
        self.assertTrue(all(not instrument.active for instrument in instruments))

        otp_record = EmailOTP.objects.filter(
            email=email,
            purpose=EmailOTP.Purpose.SIGNUP,
        ).first()
        self.assertTrue(otp_record.is_used)

    def test_request_otp_rejects_existing_username(self):
        User.objects.create_user(
            username="existing",
            email="existing@example.com",
            password="Secret@123",
        )

        response = self.client.post(
            self.request_otp_url,
            {"email": "new@example.com", "username": "existing"},
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("username", response.json())

    def test_login(self):
        user = User.objects.create_user(username="existing", email="existing@example.com", password="Secret@123")
        UserProfile.objects.create(
            user=user,
            brokerage_user_id="AB1234",
            api_key="APIKEY",
            mobile_number="9876543210",
        )

        response = self.client.post(
            self.login_url,
            {"username": "existing", "password": "Secret@123"},
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        data = response.json()
        self.assertEqual(data["username"], "existing")
        self.assertFalse(data["is_superuser"])
        self.assertFalse(data["is_staff"])

    def test_password_reset_flow(self):
        email = "existing@example.com"
        user = User.objects.create_user(
            username="existing",
            email=email,
            password="Secret@123",
        )

        response = self.client.post(
            self.password_request_url,
            {"email": email},
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(mail.outbox), 1)
        otp = self._extract_otp(mail.outbox[0].body)

        verify_response = self.client.post(
            self.password_verify_url,
            {"email": email, "otp": otp},
            format="json",
        )
        self.assertEqual(verify_response.status_code, status.HTTP_200_OK)

        reset_payload = {
            "email": email,
            "otp": otp,
            "password": "NewSecret@123",
            "confirm_password": "NewSecret@123",
        }
        reset_response = self.client.post(
            self.password_reset_url,
            reset_payload,
            format="json",
        )
        self.assertEqual(reset_response.status_code, status.HTTP_200_OK)

        user.refresh_from_db()
        self.assertTrue(user.check_password("NewSecret@123"))

        otp_record = EmailOTP.objects.filter(
            email=email,
            purpose=EmailOTP.Purpose.PASSWORD_RESET,
        ).order_by("-created_at").first()
        self.assertTrue(otp_record.is_used)


class HomeAPITestCase(APITestCase):
    def setUp(self):
        self.status_url = reverse("home-status")
        self.connect_url = reverse("home-connect")
        self.user = User.objects.create_user(
            username="existing",
            email="existing@example.com",
            password="Secret@123",
        )
        self.profile = UserProfile.objects.create(
            user=self.user,
            brokerage_user_id="AB1234",
            api_key="SKY1234567890",
            mobile_number="9876543210",
        )

    def test_home_status_returns_profile_details(self):
        response = self.client.get(self.status_url, {"username": "existing"})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        data = response.json()
        self.assertEqual(data["client_id"], "AB1234")
        self.assertEqual(data["connection_state"], "idle")
        self.assertIsNone(data["last_connected_at"])
        self.assertNotEqual(data["api_key_masked"], self.profile.api_key)
        self.assertTrue(data["api_key_masked"].startswith("SKY1"))

    @patch("api.views.get_profile")
    def test_home_status_validates_active_token(self, get_profile_mock):
        self.profile.jwt_token = "jwt-token"
        self.profile.save(update_fields=["jwt_token"])

        get_profile_mock.return_value = {"status": True, "message": "SUCCESS"}

        response = self.client.get(self.status_url, {"username": "existing"})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        data = response.json()

        self.assertEqual(data["connection_state"], "connected")
        self.assertIsNone(data["connection_message"])
        get_profile_mock.assert_called_once_with(
            api_key="SKY1234567890",
            jwt_token="jwt-token",
        )

        self.profile.refresh_from_db()
        self.assertEqual(self.profile.last_token_status, "success")
        self.assertNotEqual(self.profile.token_state, "invalid")

    @patch("api.views.get_profile", side_effect=AngelAPIError("Token expired"))
    def test_home_status_marks_failed_when_validation_fails(self, get_profile_mock):
        self.profile.jwt_token = "expired-token"
        self.profile.save(update_fields=["jwt_token"])

        response = self.client.get(self.status_url, {"username": "existing"})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        data = response.json()

        self.assertEqual(data["connection_state"], "failed")
        self.assertEqual(data["connection_message"], "Token expired")
        get_profile_mock.assert_called_once_with(
            api_key="SKY1234567890",
            jwt_token="expired-token",
        )

        self.profile.refresh_from_db()
        self.assertEqual(self.profile.last_token_status, "failed")
        self.assertEqual(self.profile.token_state, "invalid")

    @patch("api.views.angel_login")
    def test_home_connect_returns_connection_payload(self, angel_login_mock):
        angel_login_mock.return_value = {
            "status": True,
            "message": "success",
            "data": {
                "jwtToken": "jwt-token",
                "refreshToken": "refresh-token",
                "feedToken": "feed-token",
                "tokenType": "full",
            },
        }

        payload = {"username": "existing", "mpin": "1234", "totp": "123456"}
        response = self.client.post(self.connect_url, payload, format="json")
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        angel_login_mock.assert_called_once_with(
            api_key="SKY1234567890",
            clientcode="AB1234",
            password="1234",
            totp="123456",
        )

        data = response.json()
        self.assertIn("message", data)
        self.assertIn("last_connected_at", data)
        self.assertIn("details", data)
        self.assertEqual(data["details"]["client_id"], "AB1234")
        self.assertEqual(data["details"]["connection_state"], "connected")
        self.assertIsNotNone(data["details"]["last_connected_at"])

        self.profile.refresh_from_db()
        self.assertEqual(self.profile.jwt_token, "jwt-token")
        self.assertEqual(self.profile.refresh_token, "refresh-token")
        self.assertEqual(self.profile.feed_token, "feed-token")
        self.assertEqual(self.profile.token_state, "full")

    @patch("api.views.angel_login", side_effect=AngelAPIError("Angel error"))
    def test_home_connect_handles_angel_error(self, angel_login_mock):
        payload = {"username": "existing", "mpin": "1234", "totp": "123456"}
        response = self.client.post(self.connect_url, payload, format="json")
        self.assertEqual(response.status_code, status.HTTP_502_BAD_GATEWAY)
        self.assertIn("detail", response.json())
        angel_login_mock.assert_called_once()


class StrategyAlphaIntegrationTestCase(APITestCase):
    class SequenceProvider(BaseMarketDataProvider):
        def __init__(self, prices):
            self._prices = list(prices)
            self._last = prices[-1] if prices else Decimal("0")

        def get_price(self, instrument):
            if self._prices:
                self._last = self._prices.pop(0)
            return self._last

    def setUp(self):
        self.user = User.objects.create_user(
            username="strategyuser",
            email="strategy@example.com",
            password="Secret@123",
        )
        self.config = AlgoConfiguration.objects.create(
            user=self.user,
            algo_active=True,
            market_active=True,
        )
        self.instrument = Instrument.objects.create(
            user=self.user,
            instrument=Instrument.InstrumentCode.NIFTY,
            transaction=Instrument.Transaction.BUY,
            lot_size=75,
            no_of_lots=1,
            pl_exit_lots=1,
            premium_price=Decimal("100"),
            pl_points=10,
            sl_points=5,
            trailing_points=0,
            active=True,
        )
        self.activation = StrategyActivation.objects.create(
            user=self.user,
            strategy_code=StrategyActivation.STRATEGY_ALPHA,
            is_active=True,
            execution_mode=StrategyActivation.ExecutionMode.DEMO,
        )
        self.activation.selected_instruments.add(self.instrument)
        self.pnl_url = reverse("pnl-list")
        self.export_url = reverse("pnl-export")
        self.run_url = reverse("strategy-alpha-run")

    def test_strategy_alpha_demo_cycle_generates_trades_and_pnl(self):
        with patch(
            "api.services.strategy_alpha.build_market_data_provider",
            return_value=self.SequenceProvider([Decimal("95.00")]),
        ):
            summary_open = StrategyAlphaEngine(user=self.user).run()
        self.assertEqual(summary_open["opened_trades"], 1)
        trade = Trade.objects.get()
        self.assertEqual(trade.status, Trade.Status.OPEN)

        with patch(
            "api.services.strategy_alpha.build_market_data_provider",
            return_value=self.SequenceProvider([Decimal("112.00")]),
        ):
            summary_close = StrategyAlphaEngine(user=self.user).run()

        trade.refresh_from_db()
        self.assertEqual(summary_close["closed_trades"], 1)
        self.assertEqual(trade.status, Trade.Status.CLOSED)
        self.assertGreater(trade.pnl, Decimal("0"))

        response = self.client.get(self.pnl_url, {"username": "strategyuser"})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        data = response.json()
        self.assertEqual(data["total_records"], 1)
        self.assertIn("total_profit", data)

        export_response = self.client.get(
            self.export_url,
            {"username": "strategyuser", "mode": "demo"},
        )
        self.assertEqual(export_response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            export_response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(export_response.content))
        sheet = workbook.active
        self.assertGreaterEqual(sheet.max_row, 2)

    def test_strategy_alpha_run_view(self):
        with patch(
            "api.services.strategy_alpha.build_market_data_provider",
            return_value=self.SequenceProvider([Decimal("95.00")]),
        ):
            response = self.client.post(
                self.run_url,
                {"username": "strategyuser", "mode": "demo"},
                format="json",
            )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        payload = response.json()
        self.assertIn("status", payload)
        self.assertIn(payload["status"], {"completed", "skipped"})

    def test_live_mode_requires_trading_metadata(self):
        self.activation.execution_mode = StrategyActivation.ExecutionMode.LIVE
        self.activation.save(update_fields=["execution_mode"])

        UserProfile.objects.create(
            user=self.user,
            brokerage_user_id="AB1234",
            api_key="APIKEY123",
            mobile_number="9876543210",
            jwt_token="jwt-token",
        )

        with patch(
            "api.services.strategy_alpha.build_market_data_provider",
            return_value=self.SequenceProvider([Decimal("95.00")]),
        ):
            summary = StrategyAlphaEngine(user=self.user).run()

        self.assertEqual(summary["opened_trades"], 0)
        self.assertEqual(summary["closed_trades"], 0)
        self.assertEqual(summary["status"], "completed")
        instrument_summary = summary["instrument_summaries"][0]
        self.assertIn("trading symbol", instrument_summary.get("message", "").lower())

    @override_settings(ANGEL_SANDBOX_ENABLED=True)
    def test_live_mode_with_sandbox_places_order(self):
        self.activation.execution_mode = StrategyActivation.ExecutionMode.LIVE
        self.activation.save(update_fields=["execution_mode"])

        profile = UserProfile.objects.create(
            user=self.user,
            brokerage_user_id="AB1234",
            api_key="APIKEY123",
            mobile_number="9876543210",
            jwt_token="jwt-token",
        )

        self.instrument.trading_symbol = "NIFTY-SANDBOX"
        self.instrument.symbol_token = "999999"
        self.instrument.exchange = "NFO"
        self.instrument.save(update_fields=["trading_symbol", "symbol_token", "exchange", "updated_at"])

        with patch(
            "api.services.strategy_alpha.build_market_data_provider",
            return_value=self.SequenceProvider([Decimal("95.00")]),
        ):
            summary = StrategyAlphaEngine(user=self.user).run()

        trade = Trade.objects.get()
        self.assertEqual(trade.execution_mode, Trade.ExecutionMode.LIVE)
        self.assertEqual(trade.status, Trade.Status.OPEN)
        self.assertTrue(str(trade.external_entry_id).startswith("SBX"))
        self.assertEqual(summary["opened_trades"], 1)

class AdminUserManagementAPITestCase(APITestCase):
    def setUp(self):
        self.url = reverse("admin-users")
        self.admin = User.objects.create_user(
            username="admin",
            email="admin@example.com",
            password="AdminPass@123",
        )
        self.admin.is_staff = True
        self.admin.save(update_fields=["is_staff"])

        self.target = User.objects.create_user(
            username="subscriber",
            email="subscriber@example.com",
            password="Secret@123",
        )
        UserProfile.objects.create(
            user=self.target,
            brokerage_user_id="CL1234",
            api_key="APIKEY123456",
            mobile_number="9876543210",
        )

    def test_admin_can_list_users(self):
        response = self.client.get(self.url, {"admin_username": "admin"})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        data = response.json()
        usernames = {user["username"] for user in data}
        self.assertIn("subscriber", usernames)
        self.assertIn("admin", usernames)

    def test_non_admin_cannot_access(self):
        response = self.client.get(self.url, {"admin_username": "subscriber"})
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("admin_username", response.json())


class InstrumentAPITestCase(APITestCase):
    def setUp(self):
        self.list_url = reverse("instrument-list")
        self.user = User.objects.create_user(
            username="trader",
            email="trader@example.com",
            password="Secret@123",
        )
        self.nifty = Instrument.objects.create(
            user=self.user,
            instrument=Instrument.InstrumentCode.NIFTY,
            transaction=Instrument.Transaction.BUY,
            lot_size=75,
            no_of_lots=1,
            pl_exit_lots=1,
            premium_price=Decimal("200"),
            pl_points=45,
            sl_points=35,
            trailing_points=15,
            active=False,
        )
        self.banknifty = Instrument.objects.create(
            user=self.user,
            instrument=Instrument.InstrumentCode.BANKNIFTY,
            transaction=Instrument.Transaction.BUY,
            lot_size=35,
            no_of_lots=1,
            pl_exit_lots=1,
            premium_price=Decimal("500"),
            pl_points=50,
            sl_points=50,
            trailing_points=10,
            active=False,
        )
        self.sensex = Instrument.objects.create(
            user=self.user,
            instrument=Instrument.InstrumentCode.SENSEX,
            transaction=Instrument.Transaction.BUY,
            lot_size=20,
            no_of_lots=1,
            pl_exit_lots=1,
            premium_price=Decimal("400"),
            pl_points=40,
            sl_points=35,
            trailing_points=12,
            active=False,
        )

    @patch(
        "api.views.load_expiry_map",
        return_value={
            "NIFTY": ["09DEC2025", "16DEC2025"],
            "BANKNIFTY": ["30DEC2025"],
            "SENSEX": ["11DEC2025"],
        },
    )
    def test_list_instruments_returns_available_expiries(self, load_map_mock):
        response = self.client.get(self.list_url, {"username": "trader"})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        data = response.json()
        self.assertEqual(len(data), 3)
        sample = next(item for item in data if item["instrument"] == "NIFTY")
        self.assertEqual(sample["available_expiries"], ["09DEC2025", "16DEC2025"])
        self.assertEqual(sample["contract_expiry"], "")
        self.assertEqual(sample["username"], "trader")
        load_map_mock.assert_called()

    @patch(
        "api.views.load_expiry_map",
        return_value={
            "NIFTY": ["09DEC2025", "16DEC2025"],
            "BANKNIFTY": ["30DEC2025"],
            "SENSEX": ["11DEC2025"],
        },
    )
    def test_update_instrument(self, load_map_mock):
        detail_url = f"{reverse('instrument-detail', args=[self.nifty.id])}?username=trader"
        payload = {
            "contract_expiry": "09DEC2025",
            "transaction": "BUY",
            "no_of_lots": 2,
            "pl_exit_lots": 1,
            "premium_price": "200.50",
            "pl_points": 45,
            "sl_points": 35,
            "trailing_points": 15,
            "active": True,
        }

        response = self.client.put(detail_url, payload, format="json")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        body = response.json()
        self.assertEqual(body["contract_expiry"], "09DEC2025")
        self.assertEqual(body["transaction"], "BUY")
        self.assertEqual(body["available_expiries"], ["09DEC2025", "16DEC2025"])

        self.nifty.refresh_from_db()
        self.assertEqual(self.nifty.contract_expiry_code, "09DEC2025")
        self.assertEqual(self.nifty.contract_expiry, date(2025, 12, 9))
        self.assertEqual(self.nifty.premium_price, Decimal("200.50"))

        load_map_mock.assert_called()
