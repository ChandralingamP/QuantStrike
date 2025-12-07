import random
from decimal import Decimal
from io import BytesIO

from django.conf import settings
from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import User
from django.core.mail import send_mail
from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.views.decorators.csrf import csrf_exempt
from rest_framework import status, viewsets
from rest_framework.exceptions import ValidationError
from rest_framework.generics import ListAPIView
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response
from rest_framework.views import APIView

from openpyxl import Workbook

from .angel import AngelAPIError, get_profile, login as angel_login
from .models import (
    AlgoConfiguration,
    EmailOTP,
    Instrument,
    Strategy,
    StrategyActivation,
    Trade,
    UserProfile,
)
from .serializers import (
    AdminAccessSerializer,
    AdminUserDeleteSerializer,
    AdminUserSerializer,
    AdminUserToggleSerializer,
    AlgoConfigurationSerializer,
    InstrumentSerializer,
    LoginSerializer,
    OTPRequestSerializer,
    PasswordResetRequestSerializer,
    PasswordResetSerializer,
    PasswordResetVerifySerializer,
    SignupSerializer,
    StrategySerializer,
    TradeSerializer,
)
from .services.instruments import initialize_user_instruments
from .services.strategy_alpha import StrategyAlphaEngine
from .utils.instrument_data import load_expiry_map
from .utils.trade_costs import (
    calculate_margin_required,
    calculate_total_brokerage,
)


def mask_api_key(api_key: str) -> str:
    if not api_key:
        return ""
    if len(api_key) <= 4:
        return "*" * len(api_key)
    prefix = api_key[:4]
    suffix = api_key[-4:]
    hidden_length = max(len(api_key) - 8, 0)
    return f"{prefix}{'*' * hidden_length}{suffix}"


class StrategyViewSet(viewsets.ModelViewSet):
    queryset = Strategy.objects.all()
    serializer_class = StrategySerializer


class InstrumentViewSet(viewsets.ModelViewSet):
    queryset = Instrument.objects.select_related("user").order_by("instrument")
    serializer_class = InstrumentSerializer
    http_method_names = ["get", "put", "patch", "head", "options"]

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context["expiry_map"] = load_expiry_map()
        return context

    def _resolve_username(self):
        username = self.request.query_params.get("username")
        if not username and hasattr(self.request, "data"):
            username = self.request.data.get("username")
        return (username or "").strip()

    def get_queryset(self):
        queryset = super().get_queryset()
        username = self._resolve_username()
        if username:
            queryset = queryset.filter(user__username__iexact=username)
        return queryset

    def list(self, request, *args, **kwargs):
        username = self._resolve_username()
        if not username:
            raise ValidationError({"username": "Username query parameter is required."})

        user = get_object_or_404(User, username__iexact=username)
        queryset = self.filter_queryset(self.get_queryset())

        if not queryset.exists():
            initialize_user_instruments(user)
            queryset = self.filter_queryset(self.get_queryset())

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def update(self, request, *args, **kwargs):
        username = self._resolve_username()
        if not username:
            raise ValidationError({"username": "Username query parameter is required."})
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        username = self._resolve_username()
        if not username:
            raise ValidationError({"username": "Username query parameter is required."})
        return super().partial_update(request, *args, **kwargs)


class ProfitLossPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = "page_size"
    max_page_size = 200

    def get_paginated_response(self, data):
        return Response(
            {
                "results": data,
                "page": self.page.number,
                "page_size": self.page.paginator.per_page,
                "total_pages": self.page.paginator.num_pages,
                "total_records": self.page.paginator.count,
                "total_profit": str(getattr(self, "total_profit", Decimal("0"))),
                "total_margin": str(getattr(self, "total_margin", Decimal("0"))),
                "total_brokerage": str(getattr(self, "total_brokerage", Decimal("0"))),
                "total_net_profit": str(getattr(self, "total_net_profit", Decimal("0"))),
                "mode": getattr(self, "mode", "all"),
                "next": self.get_next_link(),
                "previous": self.get_previous_link(),
            }
        )


class ProfitLossMixin:
    def _resolve_user(self, request) -> User:
        username = request.query_params.get("username") or request.data.get("username")
        if not username:
            raise ValidationError({"username": "Username is required."})
        try:
            return User.objects.get(username__iexact=username)
        except User.DoesNotExist as exc:
            raise ValidationError({"username": "User not found."}) from exc

    def _filter_queryset(self, request, user: User):
        queryset = (
            Trade.objects.filter(user=user, status__in=[Trade.Status.CLOSED, Trade.Status.OPEN])
            .select_related("instrument")
            .order_by("-exit_datetime", "-entry_datetime")
        )

        mode = request.query_params.get("mode")
        if mode:
            valid_modes = {choice for choice, _ in Trade.ExecutionMode.choices}
            if mode not in valid_modes:
                raise ValidationError({"mode": "Invalid execution mode."})
            queryset = queryset.filter(execution_mode=mode)

        date_from = request.query_params.get("date_from")
        if date_from:
            parsed = parse_date(date_from)
            if not parsed:
                raise ValidationError({"date_from": "Invalid date format. Use YYYY-MM-DD."})
            queryset = queryset.filter(exit_datetime__date__gte=parsed)

        date_to = request.query_params.get("date_to")
        if date_to:
            parsed = parse_date(date_to)
            if not parsed:
                raise ValidationError({"date_to": "Invalid date format. Use YYYY-MM-DD."})
            queryset = queryset.filter(exit_datetime__date__lte=parsed)

        return queryset


@csrf_exempt
class ProfitLossView(ProfitLossMixin, ListAPIView):
    authentication_classes = []
    permission_classes = []
    serializer_class = TradeSerializer
    pagination_class = ProfitLossPagination
    
    @csrf_exempt
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_queryset(self):
        user = self._resolve_user(self.request)
        self._report_user = user
        return self._filter_queryset(self.request, user)

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        paginator = self.paginator
        total_profit = queryset.aggregate(total=Sum("pnl"))
        total_profit_value = total_profit.get("total") if total_profit else Decimal("0")
        total_margin = Decimal("0")
        total_brokerage = Decimal("0")
        for trade in queryset.iterator():
            total_margin += calculate_margin_required(trade)
            total_brokerage += calculate_total_brokerage(trade)
        total_net_profit = (total_profit_value or Decimal("0")) - total_brokerage
        if paginator is not None:
            paginator.total_profit = total_profit_value or Decimal("0")
            paginator.total_margin = total_margin
            paginator.total_brokerage = total_brokerage
            paginator.total_net_profit = total_net_profit
            paginator.mode = request.query_params.get("mode") or "all"
            page = paginator.paginate_queryset(queryset, request, view=self)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return paginator.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(
            {
                "results": serializer.data,
                "page": 1,
                "page_size": len(serializer.data),
                "total_pages": 1,
                "total_records": len(serializer.data),
                "total_profit": str(total_profit_value or Decimal("0")),
                "total_margin": str(total_margin),
                "total_brokerage": str(total_brokerage),
                "total_net_profit": str(total_net_profit),
                "mode": request.query_params.get("mode") or "all",
            }
        )


@csrf_exempt
class ProfitLossExportView(ProfitLossMixin, APIView):
    authentication_classes = []
    permission_classes = []
    
    @csrf_exempt
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get(self, request, *args, **kwargs):
        user = self._resolve_user(request)
        queryset = self._filter_queryset(request, user).order_by("entry_datetime")
        if not queryset.exists():
            return Response(
                {"detail": "No trades available for export."},
                status=status.HTTP_404_NOT_FOUND,
            )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "ProfitLoss"
        headers = [
            "#",
            "Entry Date",
            "Exit Date",
            "Instrument",
            "Symbol",
            "Mode",
            "Direction",
            "Quantity",
            "Buy @",
            "Sell @",
            "Gross P/L",
            "Margin Needed",
            "Brokerage",
            "Net P/L",
            "Notes",
        ]
        sheet.append(headers)

        total_profit = Decimal("0")
        total_margin = Decimal("0")
        total_brokerage = Decimal("0")
        total_net_profit = Decimal("0")
        for index, trade in enumerate(queryset, start=1):
            entry_price = trade.entry_price
            exit_price = trade.exit_price
            pnl = trade.pnl or Decimal("0")
            margin_needed = calculate_margin_required(trade)
            brokerage = calculate_total_brokerage(trade)
            net_profit = pnl - brokerage
            total_profit += pnl
            total_margin += margin_needed
            total_brokerage += brokerage
            total_net_profit += net_profit
            sheet.append(
                [
                    index,
                    trade.entry_datetime.isoformat() if trade.entry_datetime else "",
                    trade.exit_datetime.isoformat() if trade.exit_datetime else "",
                    trade.instrument.get_instrument_display() if trade.instrument_id else "",
                    trade.contract_symbol or (trade.instrument.trading_symbol if trade.instrument_id else ""),
                    trade.get_execution_mode_display(),
                    trade.get_direction_display(),
                    trade.quantity,
                    float(entry_price) if entry_price is not None else "",
                    float(exit_price) if exit_price is not None else "",
                    float(pnl),
                    float(margin_needed),
                    float(brokerage),
                    float(net_profit),
                    trade.notes,
                ]
            )

        sheet.append([])
        sheet.append([
            "Total",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            float(total_profit),
            float(total_margin),
            float(total_brokerage),
            float(total_net_profit),
            "",
        ])

        from openpyxl.utils import get_column_letter  # local import to avoid top-level clutter

        for column_index in range(1, sheet.max_column + 1):
            column_letter = get_column_letter(column_index)
            sheet.column_dimensions[column_letter].width = 18

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        mode = request.query_params.get("mode") or "all"
        filename = f"quantstrike-pnl-{user.username}-{mode}-{timezone.now().date()}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


@csrf_exempt
class StrategyAlphaRunView(APIView):
    authentication_classes = []
    permission_classes = []
    
    @csrf_exempt
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        username = request.data.get("username") or request.query_params.get("username")
        if not username:
            raise ValidationError({"username": "Username is required."})
        try:
            user = User.objects.get(username__iexact=username)
        except User.DoesNotExist as exc:
            raise ValidationError({"username": "User not found."}) from exc

        mode = request.data.get("mode") or request.query_params.get("mode")
        if mode:
            valid_modes = {choice for choice, _ in Trade.ExecutionMode.choices}
            if mode not in valid_modes:
                raise ValidationError({"mode": "Invalid execution mode."})

        market_date_raw = request.data.get("market_date") or request.query_params.get("market_date")
        market_date = None
        if market_date_raw:
            market_date = parse_date(str(market_date_raw))
            if not market_date:
                raise ValidationError({"market_date": "Invalid date format. Use YYYY-MM-DD."})

        engine = StrategyAlphaEngine(user=user, execution_mode=mode, market_date=market_date)
        summary = engine.run()
        return Response(summary)


class RequestOTPView(APIView):
    authentication_classes = []
    permission_classes = []
    
    @csrf_exempt
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        serializer = OTPRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        email = serializer.validated_data["email"].lower()

        if User.objects.filter(email__iexact=email).exists():
            return Response(
                {"detail": "Email already registered. Please log in."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        otp_value = f"{random.randint(0, 999999):06d}"
        otp_hash = make_password(otp_value)
        EmailOTP.objects.filter(
            email=email,
            purpose=EmailOTP.Purpose.SIGNUP,
            is_used=False,
        ).update(is_used=True)
        EmailOTP.create_for_email(
            email=email,
            otp_hash=otp_hash,
            purpose=EmailOTP.Purpose.SIGNUP,
        )

        subject = "QuantStrike Signup OTP"
        message = (
            "Use the following One-Time Password to complete your QuantStrike sign-up: "
            f"{otp_value}. This code expires in 10 minutes."
        )
        send_mail(
            subject,
            message,
            settings.DEFAULT_FROM_EMAIL,
            [email],
            fail_silently=False,
        )

        response_data = {"detail": "OTP sent successfully."}
        if settings.DEBUG:
            response_data["otp"] = otp_value

        return Response(response_data, status=status.HTTP_200_OK)


class PasswordResetRequestView(APIView):
    authentication_classes = []
    permission_classes = []
    
    @csrf_exempt
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        serializer = PasswordResetRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        email = serializer.validated_data["email"].lower()

        user = User.objects.get(email__iexact=email)

        otp_value = f"{random.randint(0, 999999):06d}"
        otp_hash = make_password(otp_value)
        EmailOTP.objects.filter(
            email=email,
            purpose=EmailOTP.Purpose.PASSWORD_RESET,
            is_used=False,
        ).update(is_used=True)
        EmailOTP.create_for_email(
            email=email,
            otp_hash=otp_hash,
            purpose=EmailOTP.Purpose.PASSWORD_RESET,
        )

        subject = "QuantStrike Password Reset OTP"
        message = (
            "Use the following One-Time Password to reset your QuantStrike account password: "
            f"{otp_value}. This code expires in 10 minutes."
        )
        send_mail(
            subject,
            message,
            settings.DEFAULT_FROM_EMAIL,
            [email],
            fail_silently=False,
        )

        response_data = {"detail": "Password reset OTP sent successfully."}
        if settings.DEBUG:
            response_data["otp"] = otp_value

        return Response(response_data, status=status.HTTP_200_OK)


class PasswordResetVerifyView(APIView):
    authentication_classes = []
    permission_classes = []
    
    @csrf_exempt
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        serializer = PasswordResetVerifySerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        otp_record = serializer.validated_data["_otp_record"]
        return Response(
            {
                "detail": "OTP verified successfully.",
                "expires_at": otp_record.expires_at.isoformat(),
            }
        )


class PasswordResetView(APIView):
    authentication_classes = []
    permission_classes = []
    
    @csrf_exempt
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        serializer = PasswordResetSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        return Response(
            {
                "detail": "Password updated successfully.",
                "username": user.username,
            }
        )


class SignupView(APIView):
    authentication_classes = []
    permission_classes = []
    
    @csrf_exempt
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        serializer = SignupSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        return Response(
            {
                "detail": "Signup successful.",
                "username": user.username,
                "email": user.email,
            },
            status=status.HTTP_201_CREATED,
        )


class LoginView(APIView):
    authentication_classes = []
    permission_classes = []
    
    @csrf_exempt
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        serializer = LoginSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.validated_data["user"]
        return Response(
            {
                "detail": "Login successful.",
                "username": user.username,
                "email": user.email,
                "is_superuser": user.is_superuser,
                "is_staff": user.is_staff,
            }
        )


class AlgoConfigurationView(APIView):
    authentication_classes = []
    permission_classes = []
    
    @csrf_exempt
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def _resolve_user(self, request) -> User:
        username = (
            request.query_params.get("username")
            or request.data.get("username")
            if hasattr(request, "data")
            else None
        )
        if not username:
            raise ValidationError({"username": "Username is required."})
        try:
            return User.objects.get(username__iexact=username)
        except User.DoesNotExist as exc:
            raise ValidationError({"username": "User not found."}) from exc

    def _get_context(self, user: User) -> dict:
        algo_config, _ = AlgoConfiguration.objects.get_or_create(user=user)
        activation, _ = StrategyActivation.objects.get_or_create(
            user=user,
            strategy_code=StrategyActivation.STRATEGY_ALPHA,
        )
        instruments = Instrument.objects.filter(user=user).order_by("instrument")
        return {
            "algo_config": algo_config,
            "activation": activation,
            "instruments": instruments,
        }

    def get(self, request, *args, **kwargs):
        user = self._resolve_user(request)
        context = self._get_context(user)
        serializer = AlgoConfigurationSerializer(
            context={
                "user": user,
                "activation": context["activation"],
                "instrument_qs": context["instruments"],
            },
            instance=context["algo_config"],
        )
        return Response(serializer.data)

    def put(self, request, *args, **kwargs):
        user = self._resolve_user(request)
        context = self._get_context(user)
        serializer = AlgoConfigurationSerializer(
            instance=context["algo_config"],
            data=request.data,
            partial=True,
            context={
                "user": user,
                "activation": context["activation"],
                "instrument_qs": context["instruments"],
            },
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)


class HomeStatusView(APIView):
    authentication_classes = []
    permission_classes = []
    
    @csrf_exempt
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    @staticmethod
    def _validate_session(profile: UserProfile) -> tuple[bool, str | None]:
        if not profile.jwt_token:
            return False, None

        try:
            response_payload = get_profile(
                api_key=profile.api_key,
                jwt_token=profile.jwt_token,
            )
        except AngelAPIError as exc:
            message = str(exc)
        else:
            if response_payload.get("status") is False:
                message = response_payload.get("message") or "Token validation failed."
            else:
                message = response_payload.get("message")
                now = timezone.now()
                profile.last_token_status = "success"
                profile.last_token_message = (message or "")[:255]
                profile.last_token_check_at = now
                profile.save(
                    update_fields=[
                        "last_token_status",
                        "last_token_message",
                        "last_token_check_at",
                    ]
                )
                return True, None

        profile.last_token_status = "failed"
        profile.last_token_message = (message or "")[:255]
        profile.last_token_check_at = timezone.now()
        profile.token_state = "invalid"
        profile.save(
            update_fields=[
                "last_token_status",
                "last_token_message",
                "last_token_check_at",
                "token_state",
            ]
        )
        return False, message

    def get(self, request, *args, **kwargs):
        username = request.query_params.get("username")
        if not username:
            return Response(
                {"detail": "Username is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            user = User.objects.get(username__iexact=username)
        except User.DoesNotExist:
            return Response(
                {"detail": "User not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        try:
            profile = user.profile
        except UserProfile.DoesNotExist:
            profile = None

        if not profile:
            return Response(
                {
                    "client_id": None,
                    "api_key_masked": "",
                    "last_updated": None,
                    "connection_state": "idle",
                    "connection_message": None,
                    "last_connected_at": None,
                }
            )

        connection_state = "idle"
        connection_message = None

        if profile.jwt_token:
            is_valid, validation_message = self._validate_session(profile)
            if is_valid:
                connection_state = "connected"
            else:
                connection_state = "failed"
                connection_message = validation_message or "Token validation failed."
        else:
            status_value = (profile.last_token_status or "").lower()
            if status_value == "failed":
                connection_state = "failed"
                connection_message = profile.last_token_message or None

        last_connected_at = (
            profile.token_received_at.isoformat()
            if profile.token_received_at
            else None
        )

        details = {
            "client_id": profile.brokerage_user_id,
            "api_key_masked": mask_api_key(profile.api_key),
            "last_updated": profile.updated_at.isoformat(),
            "connection_state": connection_state,
            "connection_message": connection_message,
            "last_connected_at": last_connected_at,
        }
        return Response(details)


class HomeConnectView(APIView):
    authentication_classes = []
    permission_classes = []
    
    @csrf_exempt
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        username = request.data.get("username")
        mpin = request.data.get("mpin")
        totp = request.data.get("totp")

        if not username:
            return Response(
                {"detail": "Username is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not mpin or not totp:
            return Response(
                {"detail": "MPIN and TOTP are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not str(totp).isdigit() or len(str(totp)) < 4:
            return Response(
                {"detail": "Enter a valid numeric TOTP."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            user = User.objects.get(username__iexact=username)
        except User.DoesNotExist:
            return Response(
                {"detail": "User not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        try:
            profile = user.profile
        except UserProfile.DoesNotExist:
            return Response(
                {"detail": "User profile not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        try:
            login_payload = angel_login(
                api_key=profile.api_key,
                clientcode=profile.brokerage_user_id,
                password=str(mpin),
                totp=str(totp),
            )
        except AngelAPIError as exc:
            return Response(
                {"detail": str(exc)},
                status=status.HTTP_502_BAD_GATEWAY,
            )

        tokens = login_payload.get("data") or {}
        jwt_token = tokens.get("jwtToken")
        refresh_token = tokens.get("refreshToken")
        feed_token = tokens.get("feedToken", "")

        if not jwt_token or not refresh_token:
            return Response(
                {"detail": "Angel SmartAPI did not return required tokens."},
                status=status.HTTP_502_BAD_GATEWAY,
            )

        now = timezone.now()
        profile.jwt_token = jwt_token
        profile.refresh_token = refresh_token
        profile.feed_token = feed_token
        profile.token_state = tokens.get("tokenType", "connected")
        profile.token_received_at = now
        profile.last_token_check_at = now
        profile.last_token_status = "success"
        profile.last_token_message = (login_payload.get("message") or "")[:255]
        profile.save()
        profile.refresh_from_db()

        details = {
            "client_id": profile.brokerage_user_id,
            "api_key_masked": mask_api_key(profile.api_key),
            "last_updated": profile.updated_at.isoformat(),
            "connection_state": "connected",
            "connection_message": None,
            "last_connected_at": profile.token_received_at.isoformat()
            if profile.token_received_at
            else None,
        }

        return Response(
            {
                "message": "Brokerage session established successfully.",
                "last_connected_at": profile.token_received_at.isoformat()
                if profile.token_received_at
                else profile.updated_at.isoformat(),
                "details": details,
            }
        )

class AdminUserManagementView(APIView):
    authentication_classes = []
    permission_classes = []
    
    @csrf_exempt
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get(self, request, *args, **kwargs):
        serializer = AdminAccessSerializer(
            data={"admin_username": request.query_params.get("admin_username")}
        )
        serializer.is_valid(raise_exception=True)

        users = User.objects.all().order_by("username")
        return Response(AdminUserSerializer(users, many=True).data)

    def patch(self, request, *args, **kwargs):
        serializer = AdminUserToggleSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        updated_user = serializer.save()
        return Response(
            {
                "detail": "User status updated.",
                "user": AdminUserSerializer(updated_user).data,
            }
        )

    def delete(self, request, *args, **kwargs):
        serializer = AdminUserDeleteSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        deleted_username = serializer.save()
        return Response(
            {
                "detail": "User deleted.",
                "username": deleted_username,
            },
            status=status.HTTP_200_OK,
        )
